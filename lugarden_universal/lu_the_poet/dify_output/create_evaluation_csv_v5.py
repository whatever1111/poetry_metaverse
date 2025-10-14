#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
陆家明AI诗人 - 判别器D评估Excel生成器 V5
Excel + 下拉菜单版本：使用pandas + openpyxl实现人性化评估界面
"""

import os
import yaml
import re
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def extract_yaml_and_content(file_path):
    """从markdown文件中提取YAML metadata和诗歌内容"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 分离YAML和内容
        if content.startswith('---'):
            parts = content.split('---', 2)
            if len(parts) >= 3:
                yaml_content = parts[1].strip()
                poem_content = parts[2].strip()
                
                # 解析YAML
                metadata = yaml.safe_load(yaml_content)
                return metadata, poem_content
        
        return None, content
    except Exception as e:
        print(f"❌ 读取文件失败 {file_path}: {e}")
        return None, None

def extract_poem_details(poem_content, filename, metadata=None):
    """从诗歌内容和metadata中提取系列、标题、引文、纯正文"""
    lines = poem_content.split('\n')
    
    # 从文件名提取系列名
    series = ""
    if '_' in filename:
        parts = filename.split('_')
        if len(parts) >= 3:
            series = parts[2]  # 例如：250910_091743_何陋有_关于后事的清谈.md -> 何陋有
    
    # 提取标题（第一行非空行）
    title = ""
    i = 0
    while i < len(lines) and not lines[i].strip():
        i += 1
    if i < len(lines):
        title = lines[i].strip()
        i += 1
    
    # 跳过空行
    while i < len(lines) and not lines[i].strip():
        i += 1
    
    # 优先从metadata中获取引文信息（已经正确解析的）
    citation = ""
    if metadata and metadata.get('creation_meta'):
        classic_quote = metadata['creation_meta'].get('classic_quote', '')
        classic_source = metadata['creation_meta'].get('classic_source', '')
        if classic_quote and classic_source:
            citation = f"{classic_quote}\n——《{classic_source}》"
    
    # 如果metadata中没有，则尝试从内容中解析（兼容旧格式）
    if not citation:
        # 格式1: "引文内容" (带引号)
        if i < len(lines) and lines[i].strip().startswith('"'):
            quote = lines[i].strip()
            i += 1
            
            # 跳过空行
            while i < len(lines) and not lines[i].strip():
                i += 1
                
            # 提取出处（—— 开头）
            if i < len(lines) and lines[i].strip().startswith('——'):
                source = lines[i].strip()
                citation = f"{quote}\n{source}"
                i += 1
            else:
                citation = quote
        
        # 格式2: 引文内容\n——《出处》 (无引号格式，新版本)
        elif i < len(lines) and not lines[i].strip().startswith('——'):
            # 检查是否是引文行（后面跟着——《...》）
            quote_line = lines[i].strip()
            next_i = i + 1
            
            # 跳过空行找到出处行
            while next_i < len(lines) and not lines[next_i].strip():
                next_i += 1
                
            # 如果下一行是出处格式，则当前行是引文
            if (next_i < len(lines) and 
                lines[next_i].strip().startswith('——') and
                '《' in lines[next_i] and '》' in lines[next_i]):
                source = lines[next_i].strip()
                citation = f"{quote_line}\n{source}"
                i = next_i + 1
            else:
                # 不是引文格式，保持原位置
                pass
    else:
        # 如果从metadata获取了引文，需要跳过内容中的引文部分
        # 寻找引文结束位置
        while i < len(lines):
            if lines[i].strip().startswith('——') and '《' in lines[i] and '》' in lines[i]:
                i += 1
                break
            elif lines[i].strip() and not lines[i].strip().startswith('——'):
                # 可能是引文行，继续
                i += 1
            else:
                break
    
    # 跳过引文后的空行
    while i < len(lines) and not lines[i].strip():
        i += 1
    
    # 提取纯正文（剩余内容）
    poem_body_lines = []
    while i < len(lines):
        line = lines[i].strip()
        if line:  # 非空行
            poem_body_lines.append(line)
        i += 1
    
    poem_body = '\n'.join(poem_body_lines)
    
    return {
        'series': series,
        'title': title,
        'citation': citation,
        'poem_body': poem_body,  # 纯正文
        'full_poem': poem_content  # 保留完整内容
    }

def process_single_directory(dir_path, dir_name="单个目录"):
    """处理单个目录中的.md文件"""
    poems_data = []
    
    print(f"📂 处理目录: {dir_name}")
    
    # 遍历目录中的所有.md文件
    for file in os.listdir(dir_path):
        if file.endswith('.md'):
            file_path = os.path.join(dir_path, file)
            print(f"   📄 处理文件: {file}")
            
            # 提取metadata和内容
            metadata, poem_content = extract_yaml_and_content(file_path)
            if metadata and poem_content:
                
                # 提取诗歌详细信息（传递metadata）
                poem_details = extract_poem_details(poem_content, file, metadata)
                
                # 合并数据 - V5精简版（只保留关键人类评估）
                row_data = {
                    # 基本识别信息
                    '用户问题': metadata.get('creation_meta', {}).get('user_query', ''),
                    '生成轮次': metadata.get('creation_meta', {}).get('generation_round', ''),
                    
                    # 诗歌内容详细拆分
                    '系列': poem_details['series'],
                    '标题': poem_details['title'],
                    '引文': poem_details['citation'],
                    '诗歌正文': poem_details['poem_body'],  # 修复：使用纯正文
                    
                    # 判别器D评分结果
                    'D判别器_最终决策': metadata.get('evaluation', {}).get('final_decision', ''),
                    'D判别器_风格保真度': metadata.get('evaluation', {}).get('style_fidelity', ''),
                    'D判别器_引用验证': metadata.get('evaluation', {}).get('citation_verification', ''),
                    'D判别器_证据充分性': metadata.get('evaluation', {}).get('evidence_sufficiency', ''),
                    'D判别器_结构完整性': metadata.get('evaluation', {}).get('structural_integrity', ''),
                    'D判别器_错误代码': metadata.get('evaluation', {}).get('error_code', ''),
                    'D判别器_改进建议': metadata.get('evaluation', {}).get('improvement_suggestions', ''),
                    
                    # 人类评估预留列（简化版）
                    '人类评分_同意程度': ''  # 1-5分下拉菜单
                }
                
                poems_data.append(row_data)
    
    return poems_data

def scan_all_poems(base_dir="."):
    """扫描所有诗歌文件并提取信息"""
    poems_data = []
    
    # 检查是否为单个诗歌目录（直接包含带YAML metadata的.md文件）
    poetry_files_in_root = []
    for f in os.listdir(base_dir):
        if f.endswith('.md'):
            file_path = os.path.join(base_dir, f)
            metadata, _ = extract_yaml_and_content(file_path)
            if metadata:  # 只有包含YAML metadata的才算诗歌文件
                poetry_files_in_root.append(f)
    
    if poetry_files_in_root:
        # 模式1: 单个诗歌目录，直接包含.md文件
        print(f"🎯 检测到单个诗歌目录模式")
        poems_data = process_single_directory(base_dir, os.path.basename(os.path.abspath(base_dir)))
    else:
        # 模式2: 父目录，包含多个日期子目录
        print(f"🎯 检测到多目录模式，搜索日期子目录")
        
        # 遍历所有子目录
        for item in os.listdir(base_dir):
            item_path = os.path.join(base_dir, item)
            if os.path.isdir(item_path) and item.startswith('2509'):  # 只处理日期目录
                poems_data.extend(process_single_directory(item_path, item))
    
    return poems_data

def create_excel_with_dropdown(poems_data, output_file='poetry_evaluation_v5.xlsx'):
    """创建带下拉菜单的Excel文件"""
    if not poems_data:
        print("❌ 没有找到诗歌数据")
        return
    
    # V5版表头 - 简化人类评估
    fieldnames = [
        # 基本识别
        '用户问题', '生成轮次',
        
        # 诗歌内容详细拆分
        '系列', '标题', '引文', '诗歌正文',
        
        # 判别器D评分结果（评估重点）
        'D判别器_最终决策', 'D判别器_风格保真度', 'D判别器_引用验证', 
        'D判别器_证据充分性', 'D判别器_结构完整性', 'D判别器_错误代码', 'D判别器_改进建议',
        
        # 人类评估（简化版）
        '人类评分_同意程度'
    ]
    
    try:
        # Step 1: 使用pandas创建基础Excel
        df = pd.DataFrame(poems_data)
        # 确保列顺序正确
        df = df.reindex(columns=fieldnames)
        
        # 生成Excel文件
        df.to_excel(output_file, index=False, sheet_name='诗歌评估', engine='openpyxl')
        
        print(f"✅ 基础Excel文件创建成功")
        
        # Step 2: 使用openpyxl添加下拉菜单和格式优化
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        
        # 找到"人类评分_同意程度"列的位置
        human_rating_col = None
        for idx, cell in enumerate(worksheet[1], 1):  # 第一行是表头
            if cell.value == '人类评分_同意程度':
                human_rating_col = idx
                break
        
        if human_rating_col:
            # 创建下拉数据验证
            dv = DataValidation(
                type="list", 
                formula1='"1,2,3,4,5"',
                allow_blank=True
            )
            dv.error = '请选择1-5分'
            dv.errorTitle = '评分错误'
            dv.prompt = '1=完全不同意 2=不太同意 3=部分同意 4=基本同意 5=完全同意'
            dv.promptTitle = '评分说明'
            
            # 将列数字转换为字母
            col_letter = get_column_letter(human_rating_col)
            
            # 应用到该列的所有数据行（跳过表头）
            data_range = f'{col_letter}2:{col_letter}{len(df)+1}'
            worksheet.add_data_validation(dv)
            dv.add(data_range)
            
            # 设置列宽
            worksheet.column_dimensions[col_letter].width = 18
            
            # 添加表头注释
            header_cell = worksheet[f'{col_letter}1']
            comment = Comment(
                "📋 评分说明：\n\n"
                "1 = 完全不同意D判别器\n"
                "2 = 不太同意D判别器\n"
                "3 = 部分同意D判别器\n"
                "4 = 基本同意D判别器\n"
                "5 = 完全同意D判别器\n\n"
                "💡 使用方法：\n"
                "点击单元格 → 选择下拉箭头 → 选择1-5分", 
                "陆家明AI诗人评估系统"
            )
            # 设置注释框大小，让内容完整显示
            comment.width = 300  # 宽度（像素）
            comment.height = 200  # 高度（像素）
            header_cell.comment = comment
            
            print(f"✅ 下拉菜单已添加到列 {col_letter}")
        
        # 优化其他列的显示
        # 设置主要列的列宽
        col_widths = {
            'A': 20,  # 用户问题
            'B': 12,  # 生成轮次
            'C': 12,  # 系列
            'D': 20,  # 标题
            'E': 30,  # 引文
            'F': 50,  # 诗歌正文
        }
        
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 冻结表头行
        worksheet.freeze_panes = 'A2'
        
        # 设置所有单元格垂直居中对齐
        print("🎨 设置单元格格式：垂直居中...")
        center_alignment = Alignment(
            vertical='center',      # 垂直居中
            horizontal='left',      # 水平左对齐（文本）
            wrap_text=True         # 自动换行
        )
        
        # 对所有有数据的单元格应用格式
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                     min_col=1, max_col=worksheet.max_column):
            for cell in row:
                if cell.value is not None:  # 只对有内容的单元格设置格式
                    cell.alignment = center_alignment
        
        # 对数字列（如评分列）设置右对齐
        if human_rating_col:
            right_alignment = Alignment(
                vertical='center',
                horizontal='center'    # 评分列居中对齐
            )
            col_letter = get_column_letter(human_rating_col)
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet[f'{col_letter}{row}']
                cell.alignment = right_alignment
        
        print("✅ 单元格格式设置完成")
        
        # 保存文件
        workbook.save(output_file)
        
        print(f"✅ Excel文件创建成功: {output_file}")
        print(f"📊 总共包含 {len(poems_data)} 首诗歌")
        
        # 统计信息
        questions = set(row['用户问题'] for row in poems_data)
        decisions = {}
        series_count = {}
        style_scores = []
        
        for row in poems_data:
            decision = row['D判别器_最终决策']
            decisions[decision] = decisions.get(decision, 0) + 1
            
            series = row['系列']
            series_count[series] = series_count.get(series, 0) + 1
            
            # 收集风格评分
            if row['D判别器_风格保真度']:
                try:
                    style_scores.append(float(row['D判别器_风格保真度']))
                except:
                    pass
        
        print(f"📋 判别器D评估统计:")
        print(f"   🤔 问题数量: {len(questions)}")
        print(f"   📊 决策分布:")
        for decision, count in decisions.items():
            print(f"      {decision}: {count}")
        
        print(f"   🎭 系列分布:")
        for series, count in sorted(series_count.items()):
            print(f"      {series}: {count}")
        
        if style_scores:
            avg_style = sum(style_scores) / len(style_scores)
            print(f"   🎨 平均风格评分: {avg_style:.3f}")
            
    except Exception as e:
        print(f"❌ Excel创建失败: {e}")

def choose_directory():
    """选择扫描目录"""
    print("📂 请选择扫描目录:")
    print("1. 🏠 使用默认目录 (当前目录)")
    print("2. 📁 指定目录路径")
    
    while True:
        choice = input("\n请输入选择 (1-2): ").strip()
        
        if choice == "1":
            return "."
        elif choice == "2":
            print("\n📁 请输入目录路径:")
            print("💡 路径格式说明:")
            print("   - 相对路径: . (当前目录) 或 ../上级目录 或 子目录名")
            print("   - 绝对路径: C:/完整路径 或 /完整路径")
            print("   - 示例: ../other_output 或 C:/Users/Name/Desktop/poetry")
            print("   - 注意: 目录下应包含以日期命名的子目录(如250909_xxx, 250910_xxx)")
            print("   - 输出: Excel文件将保存到 ../corpus/lujiaming/ 目录")
            
            custom_path = input("\n📂 请输入路径: ").strip()
            
            if not custom_path:
                print("❌ 路径不能为空，请重新选择")
                continue
                
            # 检查目录是否存在
            if not os.path.exists(custom_path):
                print(f"❌ 目录不存在: {custom_path}")
                retry = input("是否重新输入？(y/N): ").strip().lower()
                if retry == 'y':
                    continue
                else:
                    return "."
            
            if not os.path.isdir(custom_path):
                print(f"❌ 路径不是目录: {custom_path}")
                retry = input("是否重新输入？(y/N): ").strip().lower()
                if retry == 'y':
                    continue
                else:
                    return "."
            
            print(f"✅ 已选择目录: {os.path.abspath(custom_path)}")
            return custom_path
        else:
            print("❌ 无效选择，请输入 1 或 2")

def main():
    print("🎨 陆家明AI诗人 - 判别器D评估Excel生成器 V5")
    print("============================================================")
    print("✨ 新功能：Excel + 下拉菜单，让评估更简单")
    print("📝 使用pandas + openpyxl技术栈")
    print("🎯 简化人类评估：只需1-5分评价是否同意D判别器")
    print()
    
    # 选择扫描目录
    scan_dir = choose_directory()
    
    # 扫描诗歌文件
    print(f"\n📂 扫描目录: {os.path.abspath(scan_dir)}")
    poems_data = scan_all_poems(scan_dir)
    
    if not poems_data:
        print("❌ 未找到任何诗歌文件")
        print("💡 请确保目录下包含以日期命名的子目录(如250909_xxx, 250910_xxx)")
        return
    
    # 生成时间戳
    timestamp = datetime.now().strftime('%y%m%d_%H%M')
    output_dir = "../corpus/lujiaming/"
    output_file = f'{output_dir}人类评估表_{timestamp}.xlsx'
    
    # 确保输出目录存在
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建Excel文件
    print(f"📝 创建Excel评估表: {output_file}")
    create_excel_with_dropdown(poems_data, output_file)
    
    print("\n🎯 V5版Excel评估表特性:")
    print("【智能下拉】人类评分_同意程度列包含1-5分下拉选择")
    print("【评分说明】1=完全不同意 → 5=完全同意D判别器")
    print("【用户友好】点击选择，无需记忆选项，避免输入错误")
    print("【Excel兼容】所有志愿者都能使用，支持Office/WPS等")
    print("【冻结表头】方便浏览大量数据")
    print("\n📋 志愿者使用说明:")
    print("1. 双击Excel文件打开")
    print("2. 点击'人类评分_同意程度'列的任意单元格")
    print("3. 选择下拉箭头，选择1-5分")
    print("4. 可将鼠标悬停在表头查看评分说明")

if __name__ == '__main__':
    main()
